import os
import sys
from datetime import date

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from fpdf import FPDF

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', '..'))
from db_lib import get_conn

# Screening: ROE medio dos ultimos 5 exercicios anuais (lucro do ano / patrimonio liquido no
# INICIO daquele ano, nao no fim) > 8%, soma dos lucros dos ultimos 5 anos positiva, dividendo
# medio dos ultimos 5 anos (por acao) >= 5% sobre o preco atual, liquidez corrente > 1, divida
# liquida/patrimonio liquido < 0.6 (negativo = caixa liquido, passa no filtro tambem -- e
# intencional, caixa liquido e melhor que qualquer divida positiva), patrimonio liquido > 0.
# NCAV (Ativo Circulante - Passivo Total) so e exibido, nao filtra mais (a maioria das acoes
# tem NCAV negativo, exigir positivo era restritivo demais). Ordenado por P/VP (Valor de
# Mercado / Patrimonio Liquido) ascendente.
#
# Cuidado ao ler ROE medio: e sensivel a outlier quando o patrimonio liquido no inicio de um
# dos 5 anos era muito baixo (ex: TASA3/4 tinha so R$42mi de PL no inicio de 2021, lucro de
# R$635mi naquele ano = ROE de ~1503% isolado, que sozinho puxa a media de 5 anos pra ~319%) --
# nao e bug, e a propria formula pedida reagindo a uma base de calculo pequena num ano
# especifico. Vale olhar o detalhe por ano se um ticker aparecer com ROE medio muito alto.
QUERY = '''
WITH balanco_recente AS (
    SELECT DISTINCT ON (id_ativo)
        id_ativo, dt_referencia, vl_ativo_circulante, vl_passivo_total, vl_passivo_circulante,
        vl_patrimonio_liquido, vl_divida_liquida
    FROM tb_balanco_patrimonial
    WHERE tp_periodo = 'TRIMESTRAL'
    ORDER BY id_ativo, dt_referencia DESC
),
vm_recente AS (
    SELECT DISTINCT ON (id_ativo) id_ativo, vl_valor_mercado
    FROM tb_valor_mercado
    ORDER BY id_ativo, dt_referencia DESC
),
preco_recente AS (
    SELECT DISTINCT ON (id_ativo) id_ativo, vl_fechamento, dt_cotacao
    FROM tb_cotacao
    ORDER BY id_ativo, dt_cotacao DESC
),
dre_anual AS (
    SELECT id_ativo, dt_referencia, vl_lucro_liquido,
           ROW_NUMBER() OVER (PARTITION BY id_ativo ORDER BY dt_referencia DESC) AS rn
    FROM tb_dre
    WHERE tp_periodo = 'ANUAL' AND vl_lucro_liquido IS NOT NULL
),
roe_anual AS (
    SELECT d.id_ativo, d.dt_referencia,
           d.vl_lucro_liquido / b_inicio.vl_patrimonio_liquido AS roe
    FROM dre_anual d
    JOIN tb_balanco_patrimonial b_inicio
      ON b_inicio.id_ativo = d.id_ativo
     AND b_inicio.tp_periodo = 'TRIMESTRAL'
     AND b_inicio.dt_referencia = (d.dt_referencia - INTERVAL '1 year')::date
    WHERE d.rn <= 5 AND b_inicio.vl_patrimonio_liquido > 0
),
roe_media_5anos AS (
    SELECT id_ativo, AVG(roe) AS roe_medio
    FROM roe_anual
    GROUP BY id_ativo
),
lucro_5anos AS (
    SELECT id_ativo, SUM(vl_lucro_liquido) AS lucro_total_5anos
    FROM dre_anual WHERE rn <= 5
    GROUP BY id_ativo
),
dividendos_5anos AS (
    SELECT id_ativo, SUM(vl_unitario_ajustado) / 5.0 AS dividendo_anual_medio
    FROM tb_provento
    WHERE dt_ex >= CURRENT_DATE - INTERVAL '5 years'
    GROUP BY id_ativo
)
SELECT
    a.sg_ticker,
    round((r.roe_medio * 100)::numeric, 2) AS roe_medio_pct,
    round(((dv.dividendo_anual_medio / p.vl_fechamento) * 100)::numeric, 2) AS dy_medio_pct,
    round((b.vl_ativo_circulante / b.vl_passivo_circulante)::numeric, 2) AS liquidez_corrente,
    round((b.vl_divida_liquida / b.vl_patrimonio_liquido)::numeric, 2) AS divliq_pl,
    round((b.vl_ativo_circulante - b.vl_passivo_total)::numeric, 0) AS ncav,
    round((v.vl_valor_mercado / b.vl_patrimonio_liquido)::numeric, 2) AS p_vp,
    p.dt_cotacao
FROM balanco_recente b
JOIN tb_ativo a ON a.id_ativo = b.id_ativo
JOIN vm_recente v ON v.id_ativo = b.id_ativo
JOIN preco_recente p ON p.id_ativo = b.id_ativo
JOIN roe_media_5anos r ON r.id_ativo = b.id_ativo
JOIN dividendos_5anos dv ON dv.id_ativo = b.id_ativo
JOIN lucro_5anos lc ON lc.id_ativo = b.id_ativo
WHERE b.vl_patrimonio_liquido > 0
  AND b.vl_passivo_circulante > 0
  AND (b.vl_ativo_circulante / b.vl_passivo_circulante) > 1
  AND r.roe_medio > 0.08
  AND (dv.dividendo_anual_medio / p.vl_fechamento) >= 0.05
  AND (b.vl_divida_liquida / b.vl_patrimonio_liquido) < 0.6
  AND lc.lucro_total_5anos > 0
ORDER BY p_vp ASC
'''

COLS = [
    'Ticker', 'ROE medio 5a (%)', 'DY medio 5a (%)', 'Liq. Corrente', 'DivLiq/PL',
    'NCAV', 'P/VP', 'Cotacao em',
]


def fmt_num(v):
    if v is None:
        return '-'
    return f'{v:,.2f}'.replace(',', '_').replace('.', ',').replace('_', '.')


def buscar_dados():
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(QUERY)
    linhas = cur.fetchall()
    conn.close()
    return linhas


def gerar_pdf(linhas, caminho):
    pdf = FPDF(orientation='L', unit='mm', format='A4')
    pdf.set_auto_page_break(auto=True, margin=12)
    pdf.add_page()

    pdf.set_font('Helvetica', 'B', 15)
    pdf.cell(0, 9, 'Screening: ROE, dividendo, liquidez e divida liquida/PL', new_x='LMARGIN', new_y='NEXT')
    pdf.set_font('Helvetica', '', 10)
    pdf.cell(
        0, 6,
        'Filtro: ROE medio (5 anos, PL no inicio de cada ano) > 8%, soma dos lucros de 5 anos '
        'positiva, dividendo medio 5 anos >= 5% sobre o preco atual, liquidez corrente > 1, '
        'divida liquida/PL < 0.6 (negativo = caixa liquido, passa tambem), PL > 0.',
        new_x='LMARGIN', new_y='NEXT',
    )
    pdf.cell(0, 6, 'Ordenado por P/VP (Valor de Mercado / Patrimonio Liquido) ascendente.', new_x='LMARGIN', new_y='NEXT')
    pdf.cell(0, 6, f'Gerado em {date.today().strftime("%d/%m/%Y")}', new_x='LMARGIN', new_y='NEXT')
    pdf.ln(3)

    widths = [20, 26, 24, 22, 20, 30, 18, 26]

    pdf.set_font('Helvetica', 'B', 8.5)
    pdf.set_fill_color(50, 50, 50)
    pdf.set_text_color(255, 255, 255)
    for w, c in zip(widths, COLS):
        pdf.cell(w, 8, c, border=1, align='C', fill=True)
    pdf.ln()
    pdf.set_text_color(0, 0, 0)

    pdf.set_font('Helvetica', '', 8)
    for (ticker, roe, dy, liq, divliq_pl, ncav, p_vp, dt_cot) in linhas:
        valores = [
            ticker,
            fmt_num(roe), fmt_num(dy), fmt_num(liq), fmt_num(divliq_pl),
            fmt_num(ncav), fmt_num(p_vp),
            dt_cot.strftime('%d/%m/%Y'),
        ]
        for w, v in zip(widths, valores):
            align = 'L' if v == ticker else 'R'
            pdf.cell(w, 6.5, str(v), border=1, align=align)
        pdf.ln()

    pdf.output(caminho)
    print(f'PDF gerado: {caminho}')


def gerar_excel(linhas, caminho):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Screening'

    ws.append(COLS)
    header_fill = PatternFill(start_color='323232', end_color='323232', fill_type='solid')
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for (ticker, roe, dy, liq, divliq_pl, ncav, p_vp, dt_cot) in linhas:
        ws.append([ticker, roe, dy, liq, divliq_pl, ncav, p_vp, dt_cot])

    for row in ws.iter_rows(min_row=2):
        for cell in row[1:5]:
            cell.number_format = '#,##0.00'
        cell_ncav = row[5]
        cell_ncav.number_format = '#,##0'
        row[6].number_format = '0.00'
        row[7].number_format = 'DD/MM/YYYY'

    larguras = [10, 16, 14, 12, 10, 16, 10, 12]
    for i, w in enumerate(larguras, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    wb.save(caminho)
    print(f'Excel gerado: {caminho}')


if __name__ == '__main__':
    linhas = buscar_dados()
    gerar_pdf(linhas, 'relatorio_screening.pdf')
    gerar_excel(linhas, 'relatorio_screening.xlsx')
