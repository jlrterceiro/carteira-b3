import os
import sys
from datetime import date

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from fpdf import FPDF

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', '..'))
from db_lib import get_conn

QUERY = '''
WITH balanco_recente AS (
    SELECT DISTINCT ON (id_ativo)
        id_ativo, dt_referencia, vl_ativo_circulante, vl_passivo_total,
        vl_patrimonio_liquido, vl_divida_liquida
    FROM tb_balanco_patrimonial
    ORDER BY id_ativo, dt_referencia DESC
),
vm_recente AS (
    SELECT DISTINCT ON (id_ativo) id_ativo, vl_valor_mercado
    FROM tb_valor_mercado
    ORDER BY id_ativo, dt_referencia DESC
)
SELECT
    a.sg_ticker,
    b.dt_referencia AS dt_balanco,
    round(b.vl_ativo_circulante, 2) AS ativo_circulante,
    round(b.vl_passivo_total, 2) AS passivo_total,
    round(b.vl_ativo_circulante - b.vl_passivo_total, 2) AS ncav,
    round(v.vl_valor_mercado, 2) AS valor_mercado,
    round(v.vl_valor_mercado / (b.vl_ativo_circulante - b.vl_passivo_total), 4) AS vm_sobre_ncav,
    round(b.vl_divida_liquida, 2) AS divida_liquida,
    round(b.vl_patrimonio_liquido, 2) AS patrimonio_liquido,
    round(b.vl_divida_liquida / b.vl_patrimonio_liquido, 4) AS divida_liq_sobre_pl
FROM balanco_recente b
JOIN vm_recente v ON v.id_ativo = b.id_ativo
JOIN tb_ativo a ON a.id_ativo = b.id_ativo
WHERE b.vl_patrimonio_liquido > 0
  AND b.vl_ativo_circulante > b.vl_passivo_total
ORDER BY vm_sobre_ncav ASC
LIMIT 20
'''

COLS = [
    'Ticker', 'Balanco', 'Ativo Circulante', 'Passivo Total', 'NCAV (AC - PT)',
    'Valor de Mercado', 'VM/NCAV', 'Divida Liquida', 'Patrimonio Liquido', 'DivLiq/PL',
]


def fmt_num(v):
    if v is None:
        return '-'
    return f'{v:,.2f}'.replace(',', '_').replace('.', ',').replace('_', '.')


def buscar_dados():
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(QUERY)
    linhas = cur.fetchall()
    conn.close()
    return linhas


def gerar_pdf(linhas, caminho):
    pdf = FPDF(orientation='L', unit='mm', format='A4')
    pdf.set_auto_page_break(auto=True, margin=12)
    pdf.add_page()

    pdf.set_font('Helvetica', 'B', 15)
    pdf.cell(0, 9, 'Acoes mais descontadas por NCAV (Ativo Circulante - Passivo Total)', new_x='LMARGIN', new_y='NEXT')
    pdf.set_font('Helvetica', '', 10)
    pdf.cell(0, 6, 'Filtro: patrimonio liquido > 0 e ativo circulante > passivo total. Ordenado por Valor de Mercado / NCAV (menor = mais barato).', new_x='LMARGIN', new_y='NEXT')
    pdf.cell(0, 6, f'Gerado em {date.today().strftime("%d/%m/%Y")}', new_x='LMARGIN', new_y='NEXT')
    pdf.ln(3)

    widths = [18, 22, 32, 32, 32, 32, 20, 30, 32, 22]

    pdf.set_font('Helvetica', 'B', 8.5)
    pdf.set_fill_color(50, 50, 50)
    pdf.set_text_color(255, 255, 255)
    for w, c in zip(widths, COLS):
        pdf.cell(w, 8, c, border=1, align='C', fill=True)
    pdf.ln()
    pdf.set_text_color(0, 0, 0)

    pdf.set_font('Helvetica', '', 8)
    for (ticker, dt_balanco, ac, pt, ncav, vm, ratio, dl, pl, dlpl) in linhas:
        valores = [
            ticker,
            dt_balanco.strftime('%d/%m/%Y'),
            fmt_num(ac), fmt_num(pt), fmt_num(ncav), fmt_num(vm),
            fmt_num(ratio), fmt_num(dl), fmt_num(pl), fmt_num(dlpl),
        ]
        for w, v in zip(widths, valores):
            align = 'L' if v == ticker else 'R'
            pdf.cell(w, 6.5, str(v), border=1, align=align)
        pdf.ln()

    pdf.output(caminho)
    print(f'PDF gerado: {caminho}')


def gerar_excel(linhas, caminho):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'NCAV'

    ws.append(COLS)
    header_fill = PatternFill(start_color='323232', end_color='323232', fill_type='solid')
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for (ticker, dt_balanco, ac, pt, ncav, vm, ratio, dl, pl, dlpl) in linhas:
        ws.append([ticker, dt_balanco, ac, pt, ncav, vm, ratio, dl, pl, dlpl])

    for row in ws.iter_rows(min_row=2):
        row[1].number_format = 'DD/MM/YYYY'
        for cell in row[2:6]:
            cell.number_format = '#,##0.00'
        row[6].number_format = '0.0000'
        for cell in row[7:9]:
            cell.number_format = '#,##0.00'
        row[9].number_format = '0.0000'

    larguras = [10, 12, 18, 18, 18, 18, 10, 16, 18, 12]
    for i, w in enumerate(larguras, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    wb.save(caminho)
    print(f'Excel gerado: {caminho}')


if __name__ == '__main__':
    linhas = buscar_dados()
    gerar_pdf(linhas, 'relatorio_ncav.pdf')
    gerar_excel(linhas, 'relatorio_ncav.xlsx')
